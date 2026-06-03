"""
ACH Batch CSV Generator (Streamlit) - simple version.

Upload a checks-report .xlsx, click Generate, download the ACH upload CSV.

Account and routing numbers are written EXACTLY: no scientific notation,
no dropped leading zeros, no rounding. Every identifier is read straight
from the workbook cell and rendered as a string via Decimal.
"""

import io
import csv
from decimal import Decimal, InvalidOperation
from datetime import date, timedelta

import streamlit as st
import openpyxl

# Output template -----------------------------------------------------------
HEADER = [
    "New Batch Indicator", "ACH Company ID", "ACH Company Name",
    "ACH Payment Type", "Payment Description", "Payment Date",
    "Recipient Name", "Recipient Bank ID", "Recipient Account Number",
    "Recipient Account Type", "Recipient Status", "Send or Receive",
    "Recipient Amount",
]
TRAILING_EMPTY = 3  # reference output carried 3 trailing empty columns

COMPANY_ID = "2562600396"
COMPANY_NAME = "NATIONAL CHARITY"
PAYMENT_TYPE = "CCD"

# Source column names
COL_NAME = "Account Name"
COL_ROUTING = "Routing Number"
COL_ACCOUNT = "Account Number"
COL_AMOUNT = "Amount"
COL_DESC = "Stock # Prefix"


# Exact value formatting (no float corruption, ever) ------------------------
def id_to_string(value):
    """Routing / account numbers -> exact digit string (leading zeros kept)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        d = Decimal(str(value))
        if d == d.to_integral_value():
            return str(int(d))
        return format(d.normalize(), "f")
    return str(value).strip()


def amount_to_string(value):
    """Amount -> exact string. No rounding; trailing zeros trimmed."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return ""
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("$", "").strip()
        try:
            d = Decimal(cleaned)
        except InvalidOperation:
            return value.strip()
    else:
        d = Decimal(str(value))
    s = format(d, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def format_payment_date(d: date) -> str:
    return f"{d:%m/%d/%Y}"


def pad_routing(routing: str) -> str:
    if routing.isdigit() and 0 < len(routing) < 9:
        return routing.zfill(9)
    return routing


# Workbook reading ----------------------------------------------------------
def load_records(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if not grid:
        return []

    header_idx = 0
    for i, row in enumerate(grid):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if COL_ROUTING in cells and COL_ACCOUNT in cells:
            header_idx = i
            break

    headers = [str(c).strip() if c is not None else "" for c in grid[header_idx]]
    records = []
    for row in grid[header_idx + 1:]:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue
        rec = {h: (row[j] if j < len(row) else None) for j, h in enumerate(headers) if h}
        records.append(rec)
    return records


# Build CSV -----------------------------------------------------------------
def build_csv_bytes(records):
    payment_date = format_payment_date(date.today() + timedelta(days=1))
    rows = [HEADER + [""] * TRAILING_EMPTY]
    for rec in records:
        desc = id_to_string(rec.get(COL_DESC))
        b_row = ["B", COMPANY_ID, COMPANY_NAME, PAYMENT_TYPE, desc, payment_date,
                 "", "", "", "", "", "", ""] + [""] * TRAILING_EMPTY
        r_row = ["R", "", "", "", "", "",
                 id_to_string(rec.get(COL_NAME)),
                 pad_routing(id_to_string(rec.get(COL_ROUTING))),
                 id_to_string(rec.get(COL_ACCOUNT)),
                 "Checking", "Active", "Send",
                 amount_to_string(rec.get(COL_AMOUNT))] + [""] * TRAILING_EMPTY
        rows.append(b_row)
        rows.append(r_row)

    buf = io.StringIO()
    csv.writer(buf, quoting=csv.QUOTE_MINIMAL).writerows(rows)
    return buf.getvalue().encode("utf-8"), rows


def rows_to_xlsx_bytes(rows):
    """Write an .xlsx with every cell formatted as Text so leading zeros are
    kept and big account numbers never show as scientific notation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACH"
    for r in rows:
        ws.append([("" if v is None else str(v)) for v in r])
    # Force Text format on every populated cell.
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = "@"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def mask_account(acct: str) -> str:
    if not acct:
        return ""
    return "*" * max(0, len(acct) - 4) + acct[-4:] if len(acct) > 4 else "*" * len(acct)


# UI ------------------------------------------------------------------------
st.set_page_config(page_title="ACH Batch CSV Generator", page_icon="🏦")
st.title("ACH Batch CSV Generator")
st.write("Upload your Excel file to generate a formatted ACH upload CSV.")

uploaded = st.file_uploader("Choose an Excel file (.xlsx)", type=["xlsx", "xlsm"])

if uploaded is None:
    st.stop()

try:
    records = load_records(uploaded.read())
except Exception as e:  # noqa: BLE001
    st.error(f"Could not read the file: {e}")
    st.stop()

if not records:
    st.error("No data rows found in the file.")
    st.stop()

st.success(f"File loaded: **{len(records)} rows** found.")

with st.expander("Preview source data"):
    preview = [{
        "Recipient Name": id_to_string(r.get(COL_NAME)),
        "Routing Number": pad_routing(id_to_string(r.get(COL_ROUTING))),
        "Account Number": mask_account(id_to_string(r.get(COL_ACCOUNT))),
        "Description": id_to_string(r.get(COL_DESC)),
        "Amount": amount_to_string(r.get(COL_AMOUNT)),
    } for r in records]
    st.dataframe(preview, use_container_width=True, hide_index=True)
    st.caption("Account numbers are masked here. The downloaded CSV contains full values.")

if st.button("Generate ACH file", type="primary"):
    csv_bytes, rows = build_csv_bytes(records)
    xlsx_bytes = rows_to_xlsx_bytes(rows)
    stamp = f"{date.today() + timedelta(days=1):%Y%m%d}"

    st.success(f"Generated {len(records)} payments.")
    st.download_button(
        "Download Excel (.xlsx, columns formatted as Text)",
        data=xlsx_bytes, file_name=f"ach_upload_{stamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
    st.download_button(
        "Download CSV (raw text — do not open & re-save in Excel)",
        data=csv_bytes, file_name=f"ach_upload_{stamp}.csv", mime="text/csv",
    )
    st.caption(
        "The .xlsx opens in Excel with leading zeros intact and no scientific "
        "notation. The CSV holds the same correct values as raw text; if your "
        "bank portal takes CSV, upload it directly without opening it in Excel."
    )
